import yfinance as yf
import os
from openpyxl import load_workbook
from datetime import datetime

# Define the tickers and the corresponding data keys
data_keys = {
    'longName': 1,
    'symbol': 2,
    'sector': 4,
    'industry': 5,
    'currentPrice': 6,
    'financialCurrency': 7,
    'marketCap': 8,
    'dividendYield': 9,
    'dividendRate': 10,
    'payoutRatio': 11,
    'exDividendDate': 12,
    'lastDividendValue': 13,
    'lastDividendDate': 14
}

# Load an existing workbook and select the active worksheet
wb = load_workbook('Dividend Watchlist.xlsx')
ws = wb['My Positions']

# Define the tickers to retrieve data for
tickers = ['MSFT', 'AAPL', 'TSLA', 'ITE.TO', 'SUN', 'CVX', 'NTAR.CN', 'PBR-A', 'FRO', 'XOM', 'JNJ', 'MCD', 'MKC', 'SHW', 'BRY', 'ORC', 'OPI', 'AFL', 'BPT', 'DHR', 'INTU', 'IEP', 'HIMAX', 'CIM', 'LPG', 'OXLC', 'EC', 'TWO' , 'ACP']

# Retrieve the data from Yahoo Finance and write it to the worksheet
row_number = 2
for ticker in tickers:
    ticker_info = yf.Ticker(ticker).info
    for key, col in data_keys.items():
        if key in ticker_info:
            if key == 'exDividendDate' or key == 'lastDividendDate':
                timestamp = ticker_info[key]
                if timestamp is None:
                    ws.cell(row=row_number, column=col).value = 'None'
                else:
                    date = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d')
                    ws.cell(row=row_number, column=col).value = date
            else:
                ws.cell(row=row_number, column=col).value = ticker_info[key]
        else:
            ws.cell(row=row_number, column=col).value = 'None'
    row_number += 1

# Save the workbook
wb.save('Dividend Watchlist.xlsx')

# Open the workbook in Excel
os.startfile('Dividend Watchlist.xlsx')
print('Data written to Excel file successfully.')